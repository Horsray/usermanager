#
# Enhanced User management server for Huiying Proxy
# 绘影用户管理系统
# 本文件为绘影代理用户管理系统主服务端代码，负责用户、产品、台账、审批等核心业务逻辑。
# 各部分函数均有详细注释，说明用途、流程、交互及异常处理。
# 导入标准库及所需第三方库
import os
import sys
import json
import inspect
import argparse
import requests
import redis
from flask_session import Session
from datetime import datetime, timedelta
import time
from io import BytesIO
from functools import wraps
import os
import sqlite3
import shutil
from contextlib import closing
import random

# —— 租户映射（稳定可读）——
TENANT_MAP = {
    "huiying": {"tenant_id": "700243", "db": "user-700243.db"},
    "xiangyi": {"tenant_id": "700244", "db": "user-700244.db"},
}
TENANT_ID_TO_KEY = {v["tenant_id"]: k for k, v in TENANT_MAP.items()}

def _clone_schema_only(src_db_path: str, dst_db_path: str):
    """把 src 的表结构克隆到 dst（不复制数据）。"""
    with closing(sqlite3.connect(src_db_path)) as src, closing(sqlite3.connect(dst_db_path)) as dst:
        src.row_factory = sqlite3.Row
        cur = src.execute("SELECT sql FROM sqlite_master WHERE type IN ('table','index','trigger') "
                          "AND name NOT LIKE 'sqlite_%' AND sql IS NOT NULL;")
        stmts = [r["sql"] for r in cur.fetchall()]
        with dst:  # 事务
            for s in stmts:
                dst.execute(s)

def _ensure_tenant_db(tenant_key: str):
    """确保对应租户的 DB 存在。相一第一次使用时，会从 users.db 克隆结构。"""
    info = TENANT_MAP[tenant_key]
    db_path = info["db"]
    if os.path.exists(db_path):
        return db_path
    # 相一: 从 users.db 克隆结构；如果 users.db 不在当前目录，可按你的实际路径调整
    src_db = TENANT_MAP["huiying"]["db"]
    if not os.path.exists(src_db):
        raise RuntimeError(f"模板库 {src_db} 不存在，无法为 {tenant_key} 初始化库")
    _clone_schema_only(src_db, db_path)
    return db_path

def resolve_tenant(request_json, request_args):
    """解析租户：优先 tenantId，其次 space 参数，默认 huiying。返回 (tenant_key, tenant_id, db_path)"""
    tenant_id = None
    tenant_key = None

    # 1) 请求体里的 tenantId（你已经验证过前端会发）
    if isinstance(request_json, dict):
        tenant_id = request_json.get("tenantId")

    # 2) 兼容 ?space=huiying|xiangyi
    if not tenant_id:
        space = (request_args.get("space") or "").strip().lower()
        if space in TENANT_MAP:
            tenant_key = space
            tenant_id = TENANT_MAP[space]["tenant_id"]

    # 3) 默认 huiying
    if not tenant_id:
        tenant_key = "huiying"
        tenant_id = TENANT_MAP[tenant_key]["tenant_id"]

    if not tenant_key:
        tenant_key = TENANT_ID_TO_KEY.get(tenant_id, "huiying")

    db_path = _ensure_tenant_db(tenant_key)
    return tenant_key, tenant_id, db_path

# Allow importing shared utilities from parent directory
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
from db_utils import init_db, load_users, save_users, db_lock
import db_utils as dbm
from domain_models import (
    ACCOUNT_STATUS_AGENT_STOCK,
    ACCOUNT_STATUS_DISTRIBUTOR_STOCK,
    ACCOUNT_STATUS_SOLD,
    SALE_TYPE_DIRECT,
    SALE_TYPE_DISTRIBUTION,
    ROLE_ADMIN,
    ROLE_AGENT,
    ROLE_DISTRIBUTOR,
    TRANSACTION_ADMIN_TO_AGENT,
    TRANSACTION_AGENT_PURCHASE,
    TRANSACTION_AGENT_DIRECT_SALE,
    TRANSACTION_AGENT_TO_DISTRIBUTOR,
    TRANSACTION_DISTRIBUTOR_SALE,
    TRANSACTION_LEGACY,
    normalize_ledger_records,
    record_transaction,
    update_account_state,
)

# 导入Flask及相关工具
from flask import Flask, render_template, request, redirect, url_for, session, send_file, jsonify, flash
from werkzeug.utils import secure_filename

# Excel文件处理库
from openpyxl import Workbook, load_workbook
import pandas as pd
# 变量定义
# 基础目录及数据文件路径
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
init_db()
LEDGER_FILE = os.path.join(BASE_DIR, 'ledger.json')        # 台账数据文件
PRODUCTS_FILE = os.path.join(BASE_DIR, 'products.json')    # 产品数据文件
APPLICATIONS_FILE = os.path.join(BASE_DIR, 'applications.json') # 审批数据文件
DISTRIBUTION_REQUESTS_FILE = os.path.join(BASE_DIR, 'distribution_requests.json') # 分销进货申请数据文件

# Flask应用初始化及密钥设置
app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'huiying-secret')  # 用户身份码，生产环境请设置环境变量

# 配置会话存储 - 优先使用Redis，如果Redis不可用则回退到文件系统
REDIS_URL = os.getenv('REDIS_URL', 'redis://localhost:6379/1')
app.permanent_session_lifetime = timedelta(days=5)

try:
    # 尝试连接Redis
    redis_client = redis.Redis.from_url(REDIS_URL, decode_responses=False)
    redis_client.ping()  # 测试连接
    
    # 如果Redis连接成功，使用Redis存储会话
    redis_pool = redis.ConnectionPool.from_url(REDIS_URL, decode_responses=False, max_connections=100)
    app.config['SESSION_TYPE'] = 'redis'
    app.config['SESSION_REDIS'] = redis.Redis(connection_pool=redis_pool)
    print("Redis session storage initialized successfully")
except Exception as e:
    # 如果Redis连接失败，回退到文件系统存储
    print(f"Failed to connect to Redis: {e}")
    print("Falling back to filesystem session storage")
    app.config['SESSION_TYPE'] = 'filesystem'
    app.config['SESSION_FILE_DIR'] = os.path.join(BASE_DIR, 'flask_sessions')
    # 确保会话目录存在
    if not os.path.exists(app.config['SESSION_FILE_DIR']):
        os.makedirs(app.config['SESSION_FILE_DIR'])

# 初始化Session
Session(app)

# Patch Flask-Session to avoid rare save_session bug
try:
    from flask_session.sessions import RedisSessionInterface
    if isinstance(app.session_interface, RedisSessionInterface):
        class SafeRedisSessionInterface(RedisSessionInterface):
            def save_session(self, app, session, response):
                if session is None:
                    return
                try:
                    super().save_session(app, session, response)
                except Exception as e:
                    print(f"Error saving session to Redis: {e}")
                    # 发生错误时继续，避免影响请求处理
        iface = app.session_interface
        app.session_interface = SafeRedisSessionInterface(
            iface.redis, iface.key_prefix, iface.use_signer, iface.permanent
        )
except ImportError:
    # Flask-Session 0.8.0+ has different module structure
    pass


# --- Tenant routing (non-invasive, defaults to huiying) --------------------
@app.before_request
def _apply_tenant_space():
    """Pick space from query (?space=huiying|xiangyi) or session; default huiying.
    For the chosen space, ensure its DB exists and, if db_utils exposes DB_PATH,
    point it to the tenant-specific database. This keeps all existing code paths
    (load_users/save_users/etc.) untouched for huiying and only switches path
    for other spaces like 'xiangyi'.
    """
    # 1) resolve desired space (do NOT break existing behavior)
    chosen = (request.args.get('space') or session.get('space') or '').strip().lower()
    if chosen not in TENANT_MAP:
        chosen = 'huiying'
    session['space'] = chosen  # remember selection for subsequent requests

    # 2) ensure DB for that space exists (xiangyi will be created on first use)
    try:
        db_path = _ensure_tenant_db(chosen)
    except Exception:
        # fallback to huiying and leave production path untouched
        chosen = 'huiying'
        session['space'] = 'huiying'
        db_path = TENANT_MAP['huiying']['db']

    # 3) if db_utils exposes DB_PATH, redirect it to the tenant DB (runtime only)
    try:
        if hasattr(dbm, 'DB_PATH'):
            final_path = db_path if os.path.isabs(db_path) else os.path.join(BASE_DIR, db_path)
            dbm.DB_PATH = final_path
    except Exception:
        # If db_utils doesn't provide DB_PATH, we silently skip to avoid breaking production.
        pass

@app.context_processor
def _inject_space():
    # Expose current space to templates (optional; does not change existing pages)
    return dict(current_space=session.get('space', 'huiying'))




def get_location_from_ip(ip_address, username=None):
    """根据IP地址获取地理位置信息（带 users.json 自动判断与写入）"""
    if not ip_address or not username:
        return ip_address  # 不处理匿名或空IP情况

    # 从数据库读取用户信息
    users = load_users()

    user = users.get(username, {})
    current_ip = user.get("ip_address")
    current_loc = user.get("location", "")

    # ✅ 如果 IP 未变化且 location 存在，直接返回缓存的
    if current_ip == ip_address and current_loc:
        return current_loc

    # 🌐 否则调用外部 API 查询
    try:
        resp = requests.get(f"http://ip-api.com/json/{ip_address}?lang=zh-CN", timeout=3)
        if resp.status_code == 200:
            data = resp.json()
            if data.get('status') == 'success':
                country = data.get('country', '')
                region = data.get('regionName', '')
                city = data.get('city', '')
                location = '-'.join([p for p in [country, region, city] if p])
                if location:
                    # 📝 写入用户信息并保存
                    user['ip_address'] = ip_address
                    user['location'] = location
                    users[username] = user
                    save_users(users)
                    return location
    except Exception:
        pass

    # fallback：失败时也更新IP，但保留旧location或空
    user['ip_address'] = ip_address
    users[username] = user
    save_users(users)

    return current_loc or ip_address


def get_client_ip():
    """获取客户端真实IP地址
    用途：兼容代理环境下获取用户真实IP，优先X-Forwarded-For，再X-Real-IP，否则取remote_addr。
    """
    if request.headers.get('X-Forwarded-For'):
        return request.headers.get('X-Forwarded-For').split(',')[0].strip()
    elif request.headers.get('X-Real-IP'):
        return request.headers.get('X-Real-IP')
    else:
        return request.remote_addr


def generate_user_id(users: dict) -> str:
    """
    生成唯一用户编号。编号格式为当前时间戳+三位序号，确保唯一性。
    参数:
        users: 当前所有用户字典，用于检测已存在编号。
    返回:
        新的用户编号字符串。
    """
    ts = datetime.now().strftime('%Y%m%d%H%M%S')
    seq = 1
    for u in users.values():
        uid = str(u.get('user_id', ''))
        if uid.startswith(ts):
            try:
                num = int(uid[len(ts):])
                if num >= seq:
                    seq = num + 1
            except Exception:
                continue
    return f"{ts}{seq:03d}"
def gen_username_numeric(users: dict, prefix="huiying", digits=6) -> str:
    """生成唯一用户名：prefix + 指定位数纯数字"""
    while True:
        num = random.randint(10**(digits-1), 10**digits - 1)
        uname = f"{prefix}{num}"
        if uname not in users:
            return uname

def gen_password_numeric(digits=6) -> str:
    """生成指定位数纯数字密码"""
    return str(random.randint(10**(digits-1), 10**digits - 1))

# 用户数据的读写由 db_utils 提供


def load_ledger() -> list:
    """
    加载台账记录列表，自动补全role字段。
    用途：用于收入、销售等统计与显示。
    异常：文件不存在/损坏时返回空列表。
    """
    if os.path.exists(LEDGER_FILE):
        with open(LEDGER_FILE, 'r', encoding='utf-8') as f:
            try:
                records = json.load(f).get('records', [])
                normalised = normalize_ledger_records(records)
                for r in normalised:
                    r.setdefault('role', r.get('actor_role', 'admin'))
                return normalised
            except Exception:
                return []
    return []


def save_ledger(records: list) -> None:
    """
    保存台账记录到文件。
    """
    with open(LEDGER_FILE, 'w', encoding='utf-8') as f:
        json.dump({'records': records}, f, indent=4, ensure_ascii=False)


def load_products() -> dict:
    """
    加载产品信息字典，补全缺省字段（价格、默认标志）。
    用途：产品管理与下拉选择。
    异常：文件不存在/损坏时返回空字典。
    """
    if os.path.exists(PRODUCTS_FILE):
        with open(PRODUCTS_FILE, 'r', encoding='utf-8') as f:
            try:
                products = json.load(f).get('products', {})
                for p in products.values():
                    p.setdefault('price', 0)
                    p.setdefault('default', False)
                return products
            except Exception:
                return {}
    return {}


def save_products(products: dict) -> None:
    """
    保存产品信息到文件。
    """
    with open(PRODUCTS_FILE, 'w', encoding='utf-8') as f:
        json.dump({'products': products}, f, indent=4, ensure_ascii=False)


def load_applications() -> list:
    """
    加载代理批量申请记录列表。
    用途：审批管理。
    异常：文件不存在/损坏时返回空列表。
    """
    if os.path.exists(APPLICATIONS_FILE):
        with open(APPLICATIONS_FILE, 'r', encoding='utf-8') as f:
            try:
                return json.load(f).get('apps', [])
            except Exception:
                return []
    return []


def save_applications(apps: list) -> None:
    """
    保存代理批量申请记录到文件。
    """
    with open(APPLICATIONS_FILE, 'w', encoding='utf-8') as f:
        json.dump({'apps': apps}, f, indent=4, ensure_ascii=False)


def load_distribution_requests() -> list:
    """
    加载分销进货申请记录列表。
    用途：分销进货审批管理。
    异常：文件不存在/损坏时返回空列表。
    """
    if os.path.exists(DISTRIBUTION_REQUESTS_FILE):
        with open(DISTRIBUTION_REQUESTS_FILE, 'r', encoding='utf-8') as f:
            try:
                return json.load(f).get('requests', [])
            except Exception:
                return []
    return []


def save_distribution_requests(requests: list) -> None:
    """
    保存分销进货申请记录到文件。
    """
    with open(DISTRIBUTION_REQUESTS_FILE, 'w', encoding='utf-8') as f:
        json.dump({'requests': requests}, f, indent=4, ensure_ascii=False)


@app.context_processor
def inject_counts():
    """
    模板上下文处理器：为模板提供待审批/待申请数量。
    用途：页面角标、提示等。
    """
    apps = load_applications()
    pending_admin = sum(1 for a in apps if a.get('status') == 'pending')
    pending_agent = 0
    pending_distribution = 0
    
    if session.get('agent'):
        pending_agent = sum(
            1 for a in apps
            if a.get('agent') == session.get('agent') and a.get('status') == 'pending'
        )
        
        # 统计当前代理的待审批分销进货申请数量
        distribution_requests = load_distribution_requests()
        current_agent = session.get('agent')
        pending_distribution = sum(
            1 for req in distribution_requests
            if req.get('agent') == current_agent and req.get('status') == 'pending'
        )
    
    return dict(
        pending_approve_count=pending_admin,
        pending_apply_count=pending_agent,
        pending_distribution_count=pending_distribution,
    )

from datetime import datetime, date

def compute_sold_counts_from_users(users_dict: dict) -> tuple[int, int]:
    """
    统计今日/本月售出数量：
    以用户对象满足：
      - forsale 为 False（已售）
      - 且存在 sold_at（"YYYY-MM-DD HH:MM:SS" 或 "YYYY-MM-DD" 或 ISO）为准
    返回: (today_count, month_count)
    """
    today_str = date.today().strftime("%Y-%m-%d")
    month_str = date.today().strftime("%Y-%m")
    today_cnt = 0
    month_cnt = 0

    for v in users_dict.values():
        # 必须是已售
        if v.get('forsale') in (False, 0, 'false', 'False'):
            sold_at = v.get('sold_at')
            if not sold_at:
                continue
            s = str(sold_at)
            # 直接用字符串前缀判断（兼容你全局都用 "%Y-%m-%d %H:%M:%S"）
            if s.startswith(today_str):
                today_cnt += 1
            if s.startswith(month_str):
                month_cnt += 1
    return today_cnt, month_cnt
def admin_required(f):
    """
    装饰器：要求管理员身份访问。未登录重定向到登录页。
    """
    @wraps(f)
    def wrapper(*args, **kwargs):
        # 添加调试信息，查看session内容
        print(f"[DEBUG] admin_required check - session contents: {dict(session)}")
        print(f"[DEBUG] admin flag in session: {session.get('admin')}")
        print(f"[DEBUG] Request path: {request.path}")
        
        if not session.get('admin'):
            print("[DEBUG] Admin not authorized, redirecting to login")
            return redirect(url_for('login'))
        print("[DEBUG] Admin authorized, proceeding")
        return f(*args, **kwargs)
    return wrapper

def agent_required(f):
    """
    装饰器：要求销售代理身份访问。管理员也可以访问代理页面。
    用途：允许管理员角色访问所有代理功能。
    """
    @wraps(f)
    def wrapper(*args, **kwargs):
        # 添加调试信息
        print(f"[DEBUG] agent_required check - session contents: {dict(session)}")
        print(f"[DEBUG] agent flag: {session.get('agent')}, admin flag: {session.get('admin')}")
        print(f"[DEBUG] Request path: {request.path}")
        
        # 允许代理或管理员角色访问
        if not session.get('agent') and not session.get('admin'):
            print("[DEBUG] Neither agent nor admin authorized, redirecting to login")
            return redirect(url_for('login'))
        print("[DEBUG] Agent or admin authorized, proceeding")
        return f(*args, **kwargs)
    return wrapper

def distributor_required(f):
    """
    装饰器：要求分销商身份访问。管理员也可以访问分销商页面。
    用途：允许管理员角色访问所有分销商功能。
    """
    @wraps(f)
    def wrapper(*args, **kwargs):
        # 添加调试信息
        print(f"[DEBUG] distributor_required check - session contents: {dict(session)}")
        print(f"[DEBUG] distributor flag: {session.get('distributor')}, admin flag: {session.get('admin')}")
        print(f"[DEBUG] Request path: {request.path}")
        
        # 允许分销商或管理员角色访问
        if not session.get('distributor') and not session.get('admin'):
            print("[DEBUG] Neither distributor nor admin authorized, redirecting to login")
            return redirect(url_for('login'))
        print("[DEBUG] Distributor or admin authorized, proceeding")
        return f(*args, **kwargs)
    return wrapper





@app.route('/login', methods=['GET', 'POST'])
def login():
    """
    登录页面处理（管理员与代理）。
    用途：身份验证、会话初始化、IP及地理位置记录。
    交互：读取用户数据，写入登录信息。
    异常：密码错误或用户不存在时返回错误提示。
    """
    # 清除上次登录状态，避免角色混淆
    session.pop('admin', None)
    session.pop('agent', None)
    session.pop('distributor', None)

    # Honor explicit space selection via query (?space=...) or hidden form field
    selected_space = (request.args.get('space') or request.form.get('space') or session.get('space') or '').strip().lower()
    if selected_space in TENANT_MAP:
        session['space'] = selected_space

    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        users = load_users()
        user = users.get(username)
        fail_key = f"login_fail:{username}"
        fail_info = app.config['SESSION_REDIS'].hgetall(fail_key)
        if fail_info:
            fail_info = {k.decode(): v.decode() for k, v in fail_info.items()}
        fail_count = int(fail_info.get('count', 0))
        lock_until = float(fail_info.get('lock_until', 0))

        if lock_until and time.time() < lock_until:
            return render_template('login.html', error='密码错误5次，24小时内不可继续登录')
        if user and user.get('password') == password:
            # 登录成功，记录登录时间和来源IP
            client_ip = get_client_ip()
            user['last_login'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            user['ip_address'] = client_ip
            user['location'] = get_location_from_ip(client_ip, username)
            users[username] = user
            save_users(users)
            # 使用统一的方式清理失败记录，兼容不同的会话存储
            try:
                if hasattr(app.config, 'SESSION_REDIS') and app.config['SESSION_REDIS']:
                    app.config['SESSION_REDIS'].delete(fail_key)
                else:
                    print("[DEBUG] Not using Redis, skip cleaning fail record")
            except Exception as e:
                print(f"[DEBUG] Error cleaning fail record: {e}")
            if user.get('is_admin'):
                session.permanent = True
                session['admin'] = username
                print(f"[DEBUG] Admin user {username} logged in successfully - session after setting: {dict(session)}")
                return redirect(url_for('user_list'))
            elif user.get('is_agent') and user.get('is_distributor'):
                # 处理同时拥有代理和分销角色的用户
                session.permanent = True
                session['agent'] = username
                session['distributor'] = username
                session['dual_role'] = True  # 标记为双重角色
                # 默认进入代理界面，用户可以通过角色切换功能切换到分销界面
                return redirect(url_for('agent_users'))
            elif user.get('is_agent'):
                session.permanent = True
                session['agent'] = username
                return redirect(url_for('agent_users'))
            elif user.get('is_distributor'):
                session.permanent = True
                session['distributor'] = username
                return redirect(url_for('distributor_users'))
        # 登录失败 - 使用统一的方式处理，兼容不同的会话存储
        fail_count += 1
        mapping = {'count': fail_count}
        if fail_count >= 5:
            mapping['lock_until'] = time.time() + 24 * 3600
        
        try:
            if hasattr(app.config, 'SESSION_REDIS') and app.config['SESSION_REDIS']:
                app.config['SESSION_REDIS'].hset(fail_key, mapping=mapping)
            else:
                print(f"[DEBUG] Not using Redis, skip setting fail record: {fail_count}")
        except Exception as e:
            print(f"[DEBUG] Error setting fail record: {e}")
        app.config['SESSION_REDIS'].expire(fail_key, 24 * 3600)
        msg = f"密码错误{fail_count}次" + ("，24小时内不可继续登录" if fail_count >= 5 else "")
        return render_template('login.html', error=msg)
    return render_template('login.html')


@app.route('/logout')
def logout():
    """
    登出操作，清除登录状态。
    """
    session.pop('admin', None)
    session.pop('agent', None)
    session.pop('distributor', None)
    session.pop('dual_role', None)  # 清除双重角色标记
    return redirect(url_for('login'))


@app.route('/switch-role')
def switch_role():
    """
    角色切换功能，适用于拥有双重角色的用户
    """
    if not session.get('dual_role'):
        return redirect(url_for('login'))
    
    # 获取当前角色和目标角色
    current_role = request.args.get('to', 'agent')  # 默认切换到代理
    
    if current_role == 'distributor' and session.get('distributor'):
        return redirect(url_for('distributor_users'))
    elif current_role == 'agent' and session.get('agent'):
        return redirect(url_for('agent_users'))
    else:
        return redirect(url_for('login'))


@app.route('/')
@admin_required
def index():
    """
    首页重定向到用户列表，仅管理员可访问。
    """
    return redirect(url_for('user_list'))


@app.route('/bulk')
@admin_required
def bulk_manage():
    """
    批量操作管理页面，显示最近批量创建的账户信息。
    用途：批量导出、回显等。
    交互：读取会话中的批量账户列表。
    """
    accounts = session.get('bulk_accounts')
    info = session.get('bulk_info')
    products = load_products()
    page = int(request.args.get('page', 1))
    per_page = 20
    if accounts:
        start = (page - 1) * per_page
        page_accounts = accounts[start:start + per_page]
    else:
        page_accounts = None
    return render_template(
        'bulk.html', accounts=page_accounts, products=products,
        info=info, page=page, per_page=per_page,
        total=len(accounts) if accounts else 0
    )


@app.route('/users/random')
@admin_required
def random_account():
    users = load_users()
    uname = gen_username_numeric(users, prefix="huiying", digits=6)
    pwd = gen_password_numeric(digits=6)
    return jsonify({'username': uname, 'password': pwd})


@app.route('/bulk/export')
@admin_required
def bulk_export():
    """
    导出最近一次批量创建的账户为Excel文件。
    用途：管理员批量导出分发。
    交互：从session获取账户列表，生成Excel并下载。
    """
    accounts = session.get('bulk_accounts')
    if not accounts:
        return redirect(url_for('bulk_manage'))
    wb = Workbook()
    ws = wb.active
    ws.append(['用户名', '密码'])
    for acc in accounts:
        ws.append([acc['username'], acc['password']])
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    # 兼容不同Flask版本的下载参数
    filename = 'bulk_accounts.xlsx'
    kwargs = {}
    if 'download_name' in inspect.signature(send_file).parameters:
        kwargs['download_name'] = filename
    else:
        kwargs['attachment_filename'] = filename
    response = send_file(
        bio,
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        **kwargs
    )
    return response


@app.route('/users')
@admin_required
def user_list():
    """
    用户列表页面，支持多条件筛选、分页、排序。
    用途：管理员管理所有用户。
    交互：前端参数过滤，展示用户及产品信息。
    """
    query = request.args.get('q', '')
    source = request.args.get('source', '')
    status = request.args.get('status', '')
    sale = request.args.get('sale', '')
    sort = request.args.get('sort', 'desc')
    start = request.args.get('start', '')
    end = request.args.get('end', '')
    page = int(request.args.get('page', 1))
    per_page = max(int(request.args.get('per_page', 10)), 1)
    users = load_users()
    # 多条件筛选
    if query:
        users = {k: v for k, v in users.items() if query.lower() in k.lower()}
    if source:
        users = {k: v for k, v in users.items() if v.get('source') == source}
    if status:
        flag = status == 'enabled'
        users = {k: v for k, v in users.items() if v.get('enabled', True) == flag}
    if sale:
        flag = sale == 'forsale'
        users = {k: v for k, v in users.items() if v.get('forsale', False) == flag}
    if start:
        users = {k: v for k, v in users.items() if v.get('created_at', '') >= start}
    if end:
        users = {k: v for k, v in users.items() if v.get('created_at', '') <= end}
    items = list(users.items())
    # 管理员用户优先显示
    admins = [i for i in items if i[1].get('is_admin')]
    others = [i for i in items if not i[1].get('is_admin')]
    others.sort(key=lambda x: x[1].get('created_at', ''), reverse=(sort != 'asc'))
    items = admins + others
    total = len(items)
    page_items = items[(page - 1) * per_page: page * per_page]
    products = load_products()
    return render_template(
        'users.html', users=dict(page_items), total=total,
        page=page, per_page=per_page, query=query, source=source,
        status=status, sale=sale, sort=sort, start=start, end=end,
        products=products
    )


@app.route('/users/add', methods=['POST'])
@admin_required
def add_user():
    """
    添加新用户（管理员操作）。
    用途：表单提交新用户，写入用户数据并记录台账。
    交互：重名检查、字段补全。
    """
    username = request.form.get('username')
    password = request.form.get('password')
    nickname = request.form.get('nickname', '')
    is_admin = bool(request.form.get('is_admin'))
    is_agent = bool(request.form.get('is_agent'))
    price = float(request.form.get('price') or 0)
    product = request.form.get('product', '')
    if not username or not password:
        return redirect(url_for('user_list'))
    users = load_users()
    if username in users:
        return redirect(url_for('user_list'))
    users[username] = {
        'user_id': generate_user_id(users),
        'password': password,
        'nickname': nickname,
        'is_admin': is_admin,
        'is_agent': is_agent,
        'enabled': True,
        'source': 'add',
        'price': price,
        'product': product,
        'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'last_login': None,
        'ip_address': '',
        'location': '',
        'remark': ''
    }
    save_users(users)
    # ledger
    records = load_ledger()
    record_transaction(
        records,
        transaction_type=TRANSACTION_LEGACY,
        actor=session.get('admin'),
        actor_role=ROLE_ADMIN,
        amount=price,
        quantity=1,
        product=product,
    )
    save_ledger(records)
    return redirect(url_for('user_list'))


@app.route('/users/<name>/delete', methods=['POST'])
@admin_required
def delete_user(name):
    """
    删除指定用户（管理员操作）。
    用途：用户数据清理。
    """
    users = load_users()
    if name in users:
        users.pop(name)
        save_users(users)
    return redirect(url_for('user_list'))


@app.route('/users/<name>/toggle', methods=['POST'])
def toggle_user(name):
    """
    启用/禁用用户（管理员或所属代理可操作）。
    用途：账号管控，支持AJAX和表单。
    安全：仅管理员horsray或代理本人可操作。
    """
    users = load_users()
    user = users.get(name)
    permitted = False
    if session.get('admin') == 'horsray':
        permitted = True
    elif session.get('agent') and user and user.get('owner') == session.get('agent'):
        permitted = True
    if not permitted:
        return redirect(url_for('login'))
    if user:
        user['enabled'] = not user.get('enabled', True)
        save_users(users)
        if request.is_json or request.headers.get('Content-Type') == 'application/json':
            return jsonify({'success': True, 'enabled': user['enabled']})
    if request.is_json or request.headers.get('Content-Type') == 'application/json':
        return jsonify({'success': False}), 404
    return redirect(url_for('user_list') if session.get('admin') else url_for('agent_users'))


@app.route('/sales/users/<name>/sold', methods=['POST'])
@agent_required
def mark_sold(name):
    """
    标记代理名下某账号为已售出。
    用途：代理销售记录台账。
    交互：仅代理本人且账号处于待售状态可操作。
    """
    users = load_users()
    current = session.get('agent')
    state = users.get(name, {}).get('accounting', {}) if name in users else {}
    if (
        name in users
        and state.get('owner') == current
        and state.get('status') == ACCOUNT_STATUS_AGENT_STOCK
    ):
        sold_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        update_account_state(
            users[name],
            status=ACCOUNT_STATUS_SOLD,
            sale_type=SALE_TYPE_DIRECT,
            manager=current,
            sold_at=sold_time,
        )
        users[name]['sold_by'] = current
        save_users(users)
        records = load_ledger()
        record_transaction(
            records,
            transaction_type=TRANSACTION_AGENT_DIRECT_SALE,
            actor=current,
            actor_role=ROLE_AGENT,
            amount=float(users[name].get('price', 0) or 0),
            quantity=1,
            product=users[name].get('product', ''),
            account_username=name,
            sale_type=SALE_TYPE_DIRECT,
        )
        save_ledger(records)
        if request.is_json or request.headers.get('Accept') == 'application/json':
            return jsonify({'success': True})
    if request.is_json or request.headers.get('Accept') == 'application/json':
        return jsonify({'success': False}), 404
    return redirect(url_for('agent_users'))


@app.route('/sales/batch_sold', methods=['POST'])
@agent_required
def batch_sold():
    """
    批量标记代理名下账号为已售出。
    用途：代理批量销售，台账同步记录。
    """
    names = request.form.getlist('names')
    users = load_users()
    current = session.get('agent')
    sold_any = False
    records = load_ledger()
    for name in names:
        state = users.get(name, {}).get('accounting', {}) if name in users else {}
        if (
            name in users
            and state.get('owner') == current
            and state.get('status') == ACCOUNT_STATUS_AGENT_STOCK
        ):
            sold_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            update_account_state(
                users[name],
                status=ACCOUNT_STATUS_SOLD,
                sale_type=SALE_TYPE_DIRECT,
                manager=current,
                sold_at=sold_time,
            )
            users[name]['sold_by'] = current
            record_transaction(
                records,
                transaction_type=TRANSACTION_AGENT_DIRECT_SALE,
                actor=current,
                actor_role=ROLE_AGENT,
                amount=float(users[name].get('price', 0) or 0),
                quantity=1,
                product=users[name].get('product', ''),
                account_username=name,
                sale_type=SALE_TYPE_DIRECT,
            )
            sold_any = True
    if sold_any:
        save_users(users)
        save_ledger(records)
    return redirect(url_for('agent_users'))


@app.route('/sales/export_users', methods=['POST'])
@agent_required
def agent_export_users():
    """
    代理导出选中用户信息为Excel文件。
    用途：代理导出自己名下的用户信息，包含用户名、密码、昵称、备注字段。
    """
    user_ids = request.form.getlist('user_ids')
    if not user_ids:
        return redirect(url_for('agent_users'))
    
    users = load_users()
    current_agent = session.get('agent')
    
    # 筛选出属于当前代理的用户
    selected_users = {}
    for user_id in user_ids:
        if user_id in users and users[user_id].get('owner') == current_agent:
            selected_users[user_id] = users[user_id]
    
    if not selected_users:
        return redirect(url_for('agent_users'))
    
    # 创建Excel文件
    wb = Workbook()
    ws = wb.active
    ws.title = "用户信息导出"
    
    # 设置表头
    ws.append(['用户名', '密码', '昵称', '备注'])
    
    # 添加用户数据
    for username, info in selected_users.items():
        ws.append([
            username,
            info.get('password', ''),
            info.get('nickname', ''),
            info.get('remark', '')
        ])
    
    # 生成文件
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    
    # 生成文件名（包含代理名和时间戳）
    from datetime import datetime
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'{current_agent}_用户信息_{timestamp}.xlsx'
    
    return send_file(bio, download_name=filename, as_attachment=True)


@app.route('/sales/users/<name>/update', methods=['POST'])
@agent_required
def agent_update_user(name):
    """
    代理修改自己名下账号信息（支持AJAX更新备注）。
    用途：代理自助管理。
    安全：仅限本人；支持用户名变更、密码、昵称、产品、备注等。
    """
    users = load_users()
    current = session.get('agent')
    user = users.get(name)
    if not user or user.get('owner') != current:
        if request.is_json:
            return jsonify({'success': False}), 404
        return redirect(url_for('agent_users'))
    if request.is_json:
        data = request.get_json(silent=True) or {}
        remark = data.get('remark')
        if remark is not None:
            user['remark'] = remark
            save_users(users)
            return jsonify({'success': True})
        return jsonify({'success': False}), 400
    new_name = request.form.get('username')
    password = request.form.get('password')
    nickname = request.form.get('nickname')
    enabled = bool(request.form.get('enabled'))
    product = request.form.get('product')
    remark = request.form.get('remark', '')
    if new_name and new_name != name and new_name not in users:
        users[new_name] = user
        users.pop(name)
        name = new_name
        user = users[name]
    if password:
        user['password'] = password
    if nickname is not None:
        user['nickname'] = nickname
    user['enabled'] = enabled
    if product is not None:
        user['product'] = product
    user['remark'] = remark
    save_users(users)
    return redirect(url_for('agent_users'))


@app.route('/users/batch_action', methods=['POST'])
@admin_required
def batch_action():
    """
    管理员批量操作用户（删除、启用、禁用）。
    用途：多选批量管理。
    """
    action = request.form.get('action')
    names = request.form.getlist('names')
    users = load_users()
    for name in names:
        if name not in users:
            continue
        if action == 'delete':
            users.pop(name)
        elif action == 'enable':
            users[name]['enabled'] = True
        elif action == 'disable':
            users[name]['enabled'] = False
    save_users(users)
    return redirect(url_for('user_list'))


@app.route('/sales/batch_action', methods=['POST'])
@agent_required
def agent_batch_action():
    """
    代理批量操作自己名下用户（启用、禁用、标记已售）。
    用途：代理自助多选管理。
    """
    action = request.form.get('action')
    names = request.form.getlist('names')
    users = load_users()
    current = session.get('agent')
    records = load_ledger()
    for name in names:
        state = users.get(name, {}).get('accounting', {}) if name in users else {}
        if name not in users or state.get('owner') != current:
            continue
        if action == 'enable':
            users[name]['enabled'] = True
        elif action == 'disable':
            users[name]['enabled'] = False
        elif action == 'sold' and state.get('status') == ACCOUNT_STATUS_AGENT_STOCK:
            sold_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            update_account_state(
                users[name],
                status=ACCOUNT_STATUS_SOLD,
                sale_type=SALE_TYPE_DIRECT,
                manager=current,
                sold_at=sold_time,
            )
            users[name]['sold_by'] = current
            record_transaction(
                records,
                transaction_type=TRANSACTION_AGENT_DIRECT_SALE,
                actor=current,
                actor_role=ROLE_AGENT,
                amount=float(users[name].get('price', 0) or 0),
                quantity=1,
                product=users[name].get('product', ''),
                account_username=name,
                sale_type=SALE_TYPE_DIRECT,
            )
    save_users(users)
    save_ledger(records)
    return redirect(url_for('agent_users'))



@app.route('/users/<name>/update', methods=['POST'])
@admin_required
def update_user(name):
    """
    管理员更新用户信息（支持AJAX备注更新）。
    用途：支持表单与AJAX两种方式。
    """
    users = load_users()
    if request.is_json:
        data = request.get_json(silent=True) or {}
        remark = data.get('remark')
        if name in users and remark is not None:
            users[name]['remark'] = remark
            save_users(users)
            return jsonify({'success': True})
        return jsonify({'success': False}), 404

    new_name = request.form.get('username')
    password = request.form.get('password')
    nickname = request.form.get('nickname')
    is_admin = bool(request.form.get('is_admin'))
    is_agent = bool(request.form.get('is_agent'))
    is_distributor = bool(request.form.get('is_distributor'))
    enabled = bool(request.form.get('enabled'))
    product = request.form.get('product')
    remark = request.form.get('remark', '')
    if name in users:
        user = users[name]
        if new_name and new_name != name:
            users[new_name] = user
            users.pop(name)
            name = new_name
        if password:
            users[name]['password'] = password
        if nickname is not None:
            users[name]['nickname'] = nickname
        users[name]['is_admin'] = is_admin
        users[name]['is_agent'] = is_agent
        users[name]['is_distributor'] = is_distributor
        users[name]['enabled'] = enabled
        if product is not None:
            users[name]['product'] = product
        users[name]['remark'] = remark
        save_users(users)
    return redirect(url_for('user_list'))

# --- API: 获取用户明文用户名和密码（仅限管理员或代理本人） ---
@app.route('/api/users/<name>/credentials', methods=['GET'])
def api_user_credentials(name):
    """
    返回指定用户的用户名和密码（仅管理员或账号所属代理可见）。
    - 管理员：可查看任意用户
    - 代理：仅可查看自己名下（owner=代理用户名）的用户
    返回格式：{ code: 0, data: { username: 'xxx', password: 'yyy' } } 或 { code: 1, msg: '...' }
    """
    users = load_users()
    info = users.get(name)
    if not info:
        return jsonify({'code': 1, 'msg': 'not found'}), 404

    # 权限：管理员可看全部；代理只能看自己名下
    if session.get('admin'):
        pass
    elif session.get('agent') and info.get('owner') == session.get('agent'):
        pass
    else:
        return jsonify({'code': 1, 'msg': 'unauthorized'}), 401

    return jsonify({'code': 0, 'data': {'username': name, 'password': info.get('password', '')}})


@app.route('/users/import', methods=['POST'])
@admin_required
def import_users():
    """
    批量导入用户（Excel）。
    用途：大批量账号导入。
    交互：导入成功自动写入台账。
    """
    file = request.files.get('file')
    price = float(request.form.get('price') or 0)
    product = request.form.get('product', '')
    if not file:
        return redirect(url_for('user_list'))
    filename = secure_filename(file.filename)
    if not filename:
        return redirect(url_for('user_list'))
    wb = load_workbook(file)
    ws = wb.active
    users = load_users()
    first = True
    count = 0
    for row in ws.iter_rows(values_only=True):
        if first:
            first = False
            continue
        username = str(row[0]) if row and row[0] else None
        password = str(row[1]) if row and len(row) > 1 else None
        nickname = str(row[2]) if row and len(row) > 2 else ''
        is_admin = bool(row[3]) if row and len(row) > 3 else False
        if username and password:
            users[username] = {
                'user_id': generate_user_id(users),
                'password': password,
                'nickname': nickname,
                'is_admin': is_admin,
                'enabled': True,
                'source': 'import',
                'product': product,
                'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'last_login': None,
                'price': price,
                'ip_address': '',
                'location': ''
            }
            count += 1
    save_users(users)
    if count > 0 and price > 0:
        records = load_ledger()
        records.append({
            'time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'admin': session.get('admin'),
            'role': 'admin',
            'product': product,
            'price': price,
            'count': count,
            'revenue': price * count
        })
        save_ledger(records)
    return redirect(url_for('user_list'))


@app.route('/users/export')
@admin_required
def export_users():
    """
    导出所有用户信息为Excel文件。
    用途：管理员备份、分析。
    """
    users = load_users()
    wb = Workbook()
    ws = wb.active
    ws.append([
        '用户编号', '用户名', '密码', '昵称', '是否管理员',
        '启用', '来源', '创建时间', '最后登录', '产品', 'IP地址', '位置'
    ])
    for name, info in users.items():
        ws.append([
            info.get('user_id', ''),
            name,
            info.get('password'),
            info.get('nickname'),
            info.get('is_admin'),
            info.get('enabled'),
            info.get('source'),
            info.get('created_at'),
            info.get('last_login'),
            info.get('product',''),
            info.get('ip_address', ''),
            info.get('location', '')
        ])
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return send_file(bio, download_name='users_export.xlsx', as_attachment=True)


@app.route('/users/template')
@admin_required
def download_template():
    """
    下载用户导入模板Excel文件。
    用途：批量导入格式参考。
    """
    wb = Workbook()
    ws = wb.active
    ws.append(['用户名', '密码', '昵称', '是否管理员'])
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return send_file(bio, download_name='import_template.xlsx', as_attachment=True)


@app.route('/products')
@admin_required
def products():
    """
    产品管理页面。
    用途：显示所有产品信息。
    """
    products = load_products()
    return render_template('products.html', products=products)


@app.route('/products/add', methods=['POST'])
@admin_required
def add_product():
    """
    添加新产品。
    用途：表单提交产品信息，若设置为默认则取消其它默认。
    """
    name = request.form.get('name')
    version = request.form.get('version', '')
    ptype = request.form.get('ptype', '')
    price = float(request.form.get('price') or 0)
    default = bool(request.form.get('default'))
    if not name:
        return redirect(url_for('products'))
    products = load_products()
    if default:
        for p in products.values():
            p['default'] = False
    products[name] = {
        'name': name,
        'version': version,
        'type': ptype,
        'price': price,
        'default': default
    }
    save_products(products)
    return redirect(url_for('products'))


@app.route('/products/<path:name>/delete', methods=['POST'])
@admin_required
def delete_product(name):
    """
    删除指定产品。
    用途：产品维护。
    """
    products = load_products()
    if name in products:
        products.pop(name)
        save_products(products)
    return redirect(url_for('products'))


@app.route('/products/<path:name>/default', methods=['POST'])
@admin_required
def set_default_product(name):
    """
    设置指定产品为默认产品。
    用途：下单时默认选择。
    """
    products = load_products()
    if name in products:
        for p in products.values():
            p['default'] = False
        products[name]['default'] = True
        save_products(products)
    return redirect(url_for('products'))


@app.route('/users/bulk_create', methods=['POST'])
@admin_required
def bulk_create():
    """
    批量创建随机新用户。
    用途：管理员大批量生成账号。
    交互：会话保存本次批量信息，写入台账。
    """
    count = int(request.form.get('count', 0))
    price = float(request.form.get('price') or 0)
    product = request.form.get('product', '')
    users = load_users()
    new_accounts = []
    for _ in range(count):
        uname = gen_username_numeric(users, prefix="huiying", digits=6)
        pwd = gen_password_numeric(digits=6)
        users[uname] = {
            'user_id': generate_user_id(users),
            'password': pwd,
            'nickname': '',
            'is_admin': False,
            'enabled': True,
            'source': 'batch',
            'product': product,
            'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'last_login': None,
            'price': price,
            'ip_address': '',
            'location': ''
        }
        new_accounts.append({'username': uname, 'password': pwd})
    save_users(users)
    session['bulk_accounts'] = new_accounts
    session['bulk_info'] = {
        'product': product,
        'price': price,
        'admin': session.get('admin'),
        'time': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    }
    if count > 0 and price > 0:
        records = load_ledger()
        record_transaction(
            records,
            transaction_type=TRANSACTION_LEGACY,
            actor=session.get('admin'),
            actor_role=ROLE_ADMIN,
            amount=price * count,
            quantity=count,
            product=product,
        )
        save_ledger(records)
    return redirect(url_for('bulk_manage'))


@app.route('/ledger')
@admin_required
def ledger_view():
    """
    台账页面（仅管理员）。
    用途：显示收入统计、筛选、导出。
    交互：仅统计role=admin的记录，避免重复计算。
    """
    records = load_ledger()
    product_filter = request.args.get('product', '')
    admin_filter = request.args.get('admin', '')
    start = request.args.get('start', '')
    end = request.args.get('end', '')
    
    # 分页参数
    page = int(request.args.get('page', 1))
    per_page = int(request.args.get('per_page', 20))  # 每页显示20条记录
    
    # 过滤记录
    # Only show admin role records to avoid counting agent sales
    filtered_records = [
        r for r in records
        if (r.get('actor_role') or r.get('role')) == 'admin'
        and r.get('direction', 'in') == 'in'
    ]
    if product_filter:
        filtered_records = [r for r in filtered_records if r.get('product') == product_filter]
    if admin_filter:
        filtered_records = [
            r for r in filtered_records
            if (r.get('actor') or r.get('admin')) == admin_filter
        ]
    if start:
        filtered_records = [r for r in filtered_records if r.get('time', '') >= start]
    if end:
        filtered_records = [r for r in filtered_records if r.get('time', '') <= end]
    
    # 按时间倒序排序（最新的在前面）
    filtered_records.sort(key=lambda x: x.get('time', ''), reverse=True)
    
    # 分页计算
    total_records = len(filtered_records)
    total_pages = (total_records + per_page - 1) // per_page
    start_index = (page - 1) * per_page
    end_index = start_index + per_page
    paginated_records = filtered_records[start_index:end_index]
    
    # 分页信息
    pagination = {
        'page': page,
        'per_page': per_page,
        'total': total_records,
        'total_pages': total_pages,
        'has_prev': page > 1,
        'has_next': page < total_pages,
        'prev_page': page - 1 if page > 1 else None,
        'next_page': page + 1 if page < total_pages else None,
        'start_record': start_index + 1 if total_records > 0 else 0,
        'end_record': min(end_index, total_records)
    }
    
    # 计算统计数据
    from datetime import datetime
    today = datetime.now().strftime('%Y-%m-%d')
    this_month = datetime.now().strftime('%Y-%m')
    this_year = datetime.now().strftime('%Y')
    
    # 计算各时间段收入（只统计管理员角色的记录）
    daily = sum(
        float(r.get('amount', r.get('revenue', 0)) or 0)
        for r in records
        if (r.get('actor_role') or r.get('role')) == 'admin'
        and r.get('direction', 'in') == 'in'
        and r.get('time', '').startswith(today)
    )
    monthly = sum(
        float(r.get('amount', r.get('revenue', 0)) or 0)
        for r in records
        if (r.get('actor_role') or r.get('role')) == 'admin'
        and r.get('direction', 'in') == 'in'
        and r.get('time', '').startswith(this_month)
    )
    yearly = sum(
        float(r.get('amount', r.get('revenue', 0)) or 0)
        for r in records
        if (r.get('actor_role') or r.get('role')) == 'admin'
        and r.get('direction', 'in') == 'in'
        and r.get('time', '').startswith(this_year)
    )
    total = sum(
        float(r.get('amount', r.get('revenue', 0)) or 0)
        for r in records
        if (r.get('actor_role') or r.get('role')) == 'admin'
        and r.get('direction', 'in') == 'in'
    )
    
    # 计算销售人员统计（区分申请人和管理员）
    salesperson_stats = {}
    for r in records:
        if (r.get('actor_role') or r.get('role')) == 'admin' and r.get('direction', 'in') == 'in':
            if r.get('agent'):
                salesperson = r.get('agent')
            elif r.get('counterparty'):
                salesperson = r.get('counterparty')
            else:
                salesperson = r.get('actor') or r.get('admin')

            if salesperson not in salesperson_stats:
                salesperson_stats[salesperson] = {'count': 0, 'revenue': 0.0}
            salesperson_stats[salesperson]['count'] += int(r.get('quantity', r.get('count', 0)) or 0)
            salesperson_stats[salesperson]['revenue'] += float(r.get('amount', r.get('revenue', 0)) or 0)
    
    # 按收入排序销售人员
    sorted_salesperson_stats = sorted(salesperson_stats.items(), key=lambda x: x[1]['revenue'], reverse=True)
    
    products = load_products()
    return render_template(
        'ledger.html', records=paginated_records,
        product_filter=product_filter, admin_filter=admin_filter,
        start=start, end=end, products=products,
        daily=daily, monthly=monthly, yearly=yearly, total=total,
        pagination=pagination, salesperson_stats=sorted_salesperson_stats
    )


@app.route('/sales/users')
@agent_required
def agent_users():
    """
    代理名下用户管理页面。
    用途：代理自助筛选、分页、排序、管理账号。
    """
    users = load_users()
    current = session.get('agent')
    my_users = {k: v for k, v in users.items() if v.get('owner') == current}

    query = request.args.get('q', '')
    sale = request.args.get('sale', '')
    status = request.args.get('status', '')
    sort = request.args.get('sort', 'desc')
    start = request.args.get('start', '')
    end = request.args.get('end', '')
    page = int(request.args.get('page', 1))
    per_page = max(int(request.args.get('per_page', 20)), 1)
    nick = request.args.get('nick', '').strip()  # ← 新增
    assigned_distributor = request.args.get('assigned_distributor', '')  # ← 新增分销代理筛选

    if query:
        my_users = {k: v for k, v in my_users.items() if query.lower() in k.lower()}
    if nick:  # ← 新增
        my_users = {k: v for k, v in my_users.items() if nick.lower() in (v.get('nickname') or '').lower()}
    if assigned_distributor:  # ← 新增分销代理筛选
        my_users = {k: v for k, v in my_users.items() if v.get('assigned_distributor') == assigned_distributor}
    if status:
        flag = status == 'enabled'
        my_users = {k: v for k, v in my_users.items() if v.get('enabled', True) == flag}
    if sale:
        flag = sale == 'forsale'
        my_users = {k: v for k, v in my_users.items() if v.get('forsale', False) == flag}
    if start:
        my_users = {k: v for k, v in my_users.items() if v.get('created_at', '') >= start}
    if end:
        my_users = {k: v for k, v in my_users.items() if v.get('created_at', '') <= end}
        # ← 插入：统计“未售”数量（基于当前筛选结果，且只统计当前代理名下）
    unsold_count = sum(1 for v in my_users.values() if v.get('forsale', False))
        # ✅ 新增：仅统计当前代理名下（且经过筛选）的集合
    today_sold_count, month_sold_count = compute_sold_counts_from_users(my_users)

    # 为每个用户添加类型判断
    def get_user_type(user_data):
        # 优先检查sale_type字段（分配后的状态）
        sale_type = user_data.get('sale_type')
        if sale_type == '分销售出':
            return 'distributed'  # 分销售出
        
        # 原有逻辑
        if user_data.get('forsale', False):
            return 'stock'  # 存货
        elif user_data.get('distribution_tag', False):
            return 'distributed'  # 分销售出
        else:
            return 'direct'  # 总部直销
    
    # 为用户数据添加类型信息
    for username, user_data in my_users.items():
        user_data['user_type'] = get_user_type(user_data)

    items = list(my_users.items())
    items.sort(key=lambda x: x[1].get('created_at', ''), reverse=(sort != 'asc'))
    total = len(items)
    page_items = items[(page - 1) * per_page: page * per_page]

    return render_template(
        'users.html',
        users=dict(page_items),
        total=total,
        page=page,
        per_page=per_page,
        query=query,
        source='',
        status=status,
        sale=sale,
        sort=sort,
        start=start,
        end=end,
        nick=nick,
        assigned_distributor=assigned_distributor,  # ← 新增
        unsold_count=unsold_count, 
        today_sold_count=today_sold_count,    # ✅ 新增
        month_sold_count=month_sold_count,    # ✅ 新增
        products=load_products()
    )


@app.route('/sales/ledger')
def agent_ledger():
    """
    代理/分销商销售台账页面。
    用途：显示当前代理或分销商的销售记录与统计。
    """
    # 检查权限：必须是代理或分销商
    if not session.get('agent') and not session.get('distributor'):
        return redirect(url_for('login'))
    
    product_filter = request.args.get('product', '')
    start = request.args.get('start', '')
    end = request.args.get('end', '')
    # 获取分页参数
    page = int(request.args.get('page', 1))
    per_page = 10  # 每页显示10条记录

    # 获取当前用户名（代理或分销商）
    current_user = session.get('agent') or session.get('distributor')
    
    records = [
        r for r in load_ledger()
        if (r.get('actor') or r.get('admin')) == current_user
        and (r.get('actor_role') or r.get('role')) in {'agent', 'distributor'}
    ]
    if product_filter:
        records = [r for r in records if r.get('product') == product_filter]

    def _record_time(entry: dict) -> str:
        return entry.get('time', '')

    if start:
        records = [r for r in records if _record_time(r) >= start]
    if end:
        records = [r for r in records if _record_time(r) <= end]

    records.sort(key=lambda x: _record_time(x), reverse=True)

    from datetime import datetime
    today = datetime.now().strftime('%Y-%m-%d')
    this_month = datetime.now().strftime('%Y-%m')
    this_year = datetime.now().strftime('%Y')

    daily = monthly = yearly = total = 0.0
    for r in records:
        if r.get('direction', 'in') != 'in':
            continue
        amount = float(r.get('amount', r.get('revenue', 0)) or 0)
        total += amount
        record_time = _record_time(r)
        if record_time.startswith(today):
            daily += amount
        if record_time.startswith(this_month):
            monthly += amount
        if record_time.startswith(this_year):
            yearly += amount

    total_records = len(records)
    start_index = (page - 1) * per_page
    end_index = start_index + per_page
    paginated_records = records[start_index:end_index]
    total_pages = max(1, (total_records + per_page - 1) // per_page)

    products = load_products()
    return render_template(
        'agent_ledger.html', 
        records=paginated_records,
        product_filter=product_filter, admin_filter=current_user,
        start=start, end=end, products=products,
        daily=daily, monthly=monthly, yearly=yearly, total=total,
        page=page, per_page=per_page, total_pages=total_pages, total_records=total_records,
        start_index=start_index, end_index=end_index
    )


@app.route('/sales/apply', methods=['GET', 'POST'])
@agent_required
def apply_bulk():
    """
    代理批量申请账号页面。
    用途：提交申请，显示申请历史。
    交互：写入applications.json，session标记提交成功。
    """
    products = load_products()
    if request.method == 'POST':
        count = int(request.form.get('count', 0))
        price = float(request.form.get('price') or 0)
        purchase_price = float(request.form.get('purchase_price') or 0)
        product = request.form.get('product', '')
        apps = load_applications()
        apps.append({
            'id': os.urandom(6).hex(),
            'agent': session.get('agent'),
            'count': count,
            'price': price,
            'purchase_price': purchase_price,
            'product': product,
            'status': 'pending',
            'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        })
        save_applications(apps)
        session['apply_success'] = True
        return redirect(url_for('apply_bulk'))
    
    success = session.pop('apply_success', False)
    my_apps = [a for a in load_applications() if a.get('agent') == session.get('agent')]
    
    # 分页参数
    page = int(request.args.get('page', 1))
    per_page = 10  # 每页显示10条记录
    
    # 按时间倒序排序（最新的在前面）
    my_apps.sort(key=lambda x: x.get('created_at', ''), reverse=True)
    
    # 分页计算
    total_apps = len(my_apps)
    total_pages = (total_apps + per_page - 1) // per_page
    start_index = (page - 1) * per_page
    end_index = start_index + per_page
    paginated_apps = my_apps[start_index:end_index]
    
    # 分页信息
    pagination = {
        'page': page,
        'per_page': per_page,
        'total': total_apps,
        'total_pages': total_pages,
        'has_prev': page > 1,
        'has_next': page < total_pages,
        'prev_page': page - 1 if page > 1 else None,
        'next_page': page + 1 if page < total_pages else None,
        'start_record': start_index + 1 if total_apps > 0 else 0,
        'end_record': min(end_index, total_apps)
    }
    
    return render_template(
        'bulk.html', accounts=None, products=products, info=None,
        page=page, per_page=per_page, total=total_apps, success=success, 
        apps=paginated_apps, pagination=pagination
    )


@app.route('/applications')
@admin_required
def applications_list():
    """
    管理员审批页面，显示所有代理批量申请。
    """
    apps = load_applications()
    return render_template('applications.html', apps=apps, products=load_products())


def _approve_application(app_record):
    """
    内部函数：审批通过代理批量申请，批量生成账号并写入台账。
    用途：供审批接口调用。
    """
    users = load_users()
    new_accounts = []
    now_ts = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    for _ in range(app_record['count']):
        uname = gen_username_numeric(users, prefix="huiying", digits=6)
        pwd = gen_password_numeric(digits=6)
        users[uname] = {
            'user_id': generate_user_id(users),
            'password': pwd,
            'nickname': '',
            'is_admin': False,
            'enabled': True,
            'source': 'agent',
            'product': app_record['product'],
            'created_at': now_ts,
            'last_login': None,
            'price': app_record['price'],
            'ip_address': '',
            'location': '',
        }
        update_account_state(
            users[uname],
            owner=app_record['agent'],
            manager=app_record['agent'],
            status=ACCOUNT_STATUS_AGENT_STOCK,
            sale_type=None,
            sold_at=None,
        )
        new_accounts.append({'username': uname, 'password': pwd})
    save_users(users)
    total_amount = float(app_record['price'] * app_record['count'])
    records = load_ledger()
    record_transaction(
        records,
        transaction_type=TRANSACTION_ADMIN_TO_AGENT,
        actor=session.get('admin'),
        actor_role=ROLE_ADMIN,
        amount=total_amount,
        quantity=app_record['count'],
        product=app_record['product'],
        counterparty=app_record['agent'],
    )
    record_transaction(
        records,
        transaction_type=TRANSACTION_AGENT_PURCHASE,
        actor=app_record['agent'],
        actor_role=ROLE_AGENT,
        amount=total_amount,
        quantity=app_record['count'],
        product=app_record['product'],
        counterparty=session.get('admin'),
        direction='out',
    )
    save_ledger(records)
    app_record['status'] = 'approved'


@app.route('/applications/<app_id>/approve', methods=['POST'])
@admin_required
def approve_application(app_id):
    """
    审批通过指定代理批量申请。
    用途：管理员操作。
    """
    apps = load_applications()
    app_record = next((a for a in apps if a['id'] == app_id), None)
    if not app_record or app_record['status'] != 'pending':
        return redirect(url_for('applications_list'))
    _approve_application(app_record)
    save_applications(apps)
    return redirect(url_for('applications_list'))


@app.route('/applications/<app_id>/reject', methods=['POST'])
@admin_required
def reject_application(app_id):
    """
    拒绝指定代理批量申请。
    用途：管理员操作。
    """
    apps = load_applications()
    for a in apps:
        if a['id'] == app_id and a['status'] == 'pending':
            a['status'] = 'rejected'
            break
    save_applications(apps)
    return redirect(url_for('applications_list'))


@app.route('/applications/<app_id>/update', methods=['POST'])
@admin_required
def update_application(app_id):
    """
    更新指定申请的数量、单价、产品。
    用途：管理员审批前可修正申请内容。
    """
    apps = load_applications()
    app_record = next((a for a in apps if a['id'] == app_id), None)
    if not app_record or app_record.get('status') != 'pending':
        return redirect(url_for('applications_list'))
    app_record['count'] = int(request.form.get('count', app_record['count']))
    app_record['price'] = float(request.form.get('price', app_record['price']))
    app_record['product'] = request.form.get('product', app_record['product'])
    save_applications(apps)
    return redirect(url_for('applications_list'))
@app.route('/admin/migrate/backfill_sold_at', methods=['POST'])
@admin_required
def backfill_sold_at():
    """
    一次性回填历史 sold_at：
    - 仅使用代理台账（role='agent'）作为销售发生时间来源
    - 按 agent+product 维度，把每条台账的时间分配给缺少 sold_at 的已售用户
    - 分配策略：优先创建时间最早、且尚未分配过 sold_at 的记录
    返回：执行摘要
    """
    users = load_users()
    records = [
        r for r in load_ledger()
        if r.get('role') == 'agent' and r.get('count', 0) > 0
    ]

    # 建桶：按 (agent, product) 聚合销售时间，多次出现就重复加入（根据 count）
    from collections import defaultdict, deque
    sales_times = defaultdict(list)  # key=(agent, product) -> [time, time, ...]
    for r in records:
        agent = r.get('admin')  # 你的台账里代理写在 'admin' 字段
        product = r.get('product', '')
        cnt = int(r.get('count', 0) or 0)
        t = r.get('time')
        if not (agent and t and cnt > 0):
            continue
        sales_times[(agent, product)].extend([t] * cnt)

    # 每个桶内时间按先后排序，便于从早到晚分配
    for k in sales_times:
        sales_times[k].sort()
        sales_times[k] = deque(sales_times[k])

    # 为每个 (agent, product) 找到候选用户（已售且无 sold_at），按创建时间排序
    affected = 0
    for (agent, product), times in sales_times.items():
        if not times:
            continue
        candidates = [
            (name, info) for name, info in users.items()
            if info.get('owner') == agent
               and not info.get('forsale', False)           # 已售
               and not info.get('sold_at')                  # 但没时间
               and (product == '' or info.get('product', '') == product)
        ]
        # 旧数据 created_at 可能为空，排序时做兜底
        def _key(it):
            ct = it[1].get('created_at') or ''
            return ct
        candidates.sort(key=_key)

        while times and candidates:
            t = times.popleft()
            name, info = candidates.pop(0)
            info['sold_at'] = t
            info['sold_by'] = agent
            users[name] = info
            affected += 1

    if affected:
        save_users(users)

    return jsonify({
        'ok': True,
        'updated': affected,
        'note': '仅为已售且缺少 sold_at 的用户回填时间；多余的台账数量会被忽略。'
    })

@app.route('/applications/batch', methods=['POST'])
@admin_required
def batch_applications():
    """
    批量审批或拒绝多个代理申请。
    用途：提升审批效率。
    """
    action = request.form.get('action')
    ids = request.form.getlist('ids')
    if action == 'approve':
        apps = load_applications()
        for app_record in apps:
            if app_record['id'] in ids and app_record['status'] == 'pending':
                _approve_application(app_record)
        save_applications(apps)
    elif action == 'reject':
        apps = load_applications()
        for a in apps:
            if a['id'] in ids and a['status'] == 'pending':
                a['status'] = 'rejected'
        save_applications(apps)
    return redirect(url_for('applications_list'))
# ====== 统计聚合（替换 _simple_aggregate，新增更全的 _aggregate_stats） ======
from collections import Counter
from datetime import datetime, date, timedelta

def _parse_dt(s):
    if not s:
        return None
    s = str(s)
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt)
        except Exception:
            pass
    try:
        return datetime.fromisoformat(s)
    except Exception:
        return None

def _days_list(n: int, end: date | None = None):
    end = end or date.today()
    return [end - timedelta(days=i) for i in range(n-1, -1, -1)]

def _months_list(n: int, end: date | None = None):
    end = end or date.today()
    y, m = end.year, end.month
    arr = []
    for _ in range(n):
        arr.append(f"{y:04d}-{m:02d}")
        m -= 1
        if m == 0:
            m = 12; y -= 1
    return list(reversed(arr))

def _pct_change(cur: float, prev: float) -> float:
    if prev == 0:
        return 100.0 if cur > 0 else 0.0
    return (cur - prev) / prev * 100.0

def _aggregate_stats(users: dict, ledger: list, agent_view=False):
    """生成统计数据：最近30天/12个月的 售出/收入/新增 + KPI(日/月环比) + 地域Top10"""
    today = date.today()
    day_labels = [d.strftime("%Y-%m-%d") for d in _days_list(30, today)]
    mon_labels = _months_list(12, today)

    # ---- 日维度（售出人数=按 users.sold_at）----
    sold_day = Counter()
    for info in users.values():
        sa = info.get('sold_at')
        if info.get('forsale') in (False, 0, 'false', 'False') and sa:
            dt = _parse_dt(sa)
            if dt:
                sold_day[dt.strftime("%Y-%m-%d")] += 1
    
    # 收入日维度仍用 ledger，不改
    rev_day = Counter()
    # 新增：从台账计算售出数量
    sold_count_day = Counter()
    for r in ledger:
        role = r.get('actor_role') or r.get('role')
        if role == 'agent' and r.get('direction', 'in') == 'in':
            if agent_view and r.get('transaction_type') == 'distribution_assignment':
                continue
            dt = _parse_dt(r.get('time'))
            if dt:
                day_key = dt.strftime("%Y-%m-%d")
                rev_day[day_key] += float(r.get('amount', r.get('revenue', 0)) or 0)
                sold_count_day[day_key] += int(r.get('quantity', r.get('count', 0)) or 0)

    new_day = Counter()
    for info in users.values():
        dt = _parse_dt(info.get('created_at'))
        if dt:
            new_day[dt.strftime("%Y-%m-%d")] += 1

    day_sold = [int(sold_count_day.get(d, 0)) for d in day_labels]  # 使用台账数据
    day_rev  = [float(rev_day.get(d, 0)) for d in day_labels]
    day_new  = [int(new_day.get(d, 0)) for d in day_labels]

    # ---- 月维度 ----
    sold_mon = Counter()
    rev_mon  = Counter()
    new_mon  = Counter()
    # 新增：从台账计算月度售出数量
    sold_count_mon = Counter()
    
    # 售出人数：按 users.sold_at（这一步让柱状图跟你回填后的数据同步）
    for info in users.values():
        sa = info.get('sold_at')
        if info.get('forsale') in (False, 0, 'false', 'False') and sa:
            dt = _parse_dt(sa)
            if dt:
                sold_mon[dt.strftime("%Y-%m")] += 1
    
    # 收入：仍从（已过滤过的）ledger 聚合
    for r in ledger:
        role = r.get('actor_role') or r.get('role')
        if role == 'agent' and r.get('direction', 'in') == 'in':
            if agent_view and r.get('transaction_type') == 'distribution_assignment':
                continue
            dt = _parse_dt(r.get('time'))
            if dt:
                month_key = dt.strftime("%Y-%m")
                rev_mon[month_key] += float(r.get('amount', r.get('revenue', 0)) or 0)
                sold_count_mon[month_key] += int(r.get('quantity', r.get('count', 0)) or 0)
    
    # 新增用户：保持原写法
    for info in users.values():
        dt = _parse_dt(info.get('created_at'))
        if dt:
            new_mon[dt.strftime("%Y-%m")] += 1

    mon_sold = [int(sold_count_mon.get(m, 0)) for m in mon_labels]  # 使用台账数据
    mon_rev  = [float(rev_mon.get(m, 0)) for m in mon_labels]
    mon_new  = [int(new_mon.get(m, 0)) for m in mon_labels]

    # ---- KPI & 环比 ----
    tk  = today.strftime("%Y-%m-%d")
    ytk = (today - timedelta(days=1)).strftime("%Y-%m-%d")
    mk  = today.strftime("%Y-%m")
    last_month_key = (today.replace(day=1) - timedelta(days=1)).strftime("%Y-%m")

    kpis = {
        # ✅ 今日/昨日/本月/上月售出，全部从 ledger 聚合
        "today_sold": int(sold_count_day.get(tk, 0)),  # 从台账计算售出数量
        "yesterday_sold": int(sold_count_day.get(ytk, 0)),  # 从台账计算售出数量
        "month_sold": int(sold_count_mon.get(mk, 0)),  # 从台账计算售出数量
        "last_month_sold": int(sold_count_mon.get(last_month_key, 0)),  # 从台账计算售出数量

        "today_rev": float(rev_day.get(tk, 0.0)),
        "yesterday_rev": float(rev_day.get(ytk, 0.0)),
        "month_rev": float(rev_mon.get(mk, 0.0)),
        "last_month_rev": float(rev_mon.get(last_month_key, 0.0)),

        "today_new": int(new_day.get(tk, 0)),
        "yesterday_new": int(new_day.get(ytk, 0)),
        "month_new": int(new_mon.get(mk, 0)),
        "last_month_new": int(new_mon.get(last_month_key, 0)),
    }
    kpis.update({
        "dod_sold": _pct_change(kpis["today_sold"], kpis["yesterday_sold"]),
        "mom_sold": _pct_change(kpis["month_sold"], kpis["last_month_sold"]),
        "dod_rev":  _pct_change(kpis["today_rev"],  kpis["yesterday_rev"]),
        "mom_rev":  _pct_change(kpis["month_rev"],  kpis["last_month_rev"]),
        "dod_new":  _pct_change(kpis["today_new"],  kpis["yesterday_new"]),
        "mom_new":  _pct_change(kpis["month_new"],  kpis["last_month_new"]),
    })

    # ---- 地域分布（国家 Top10 + 其它；中国则按省份 Top10 + 其它）----
    import re  # 如果文件顶部已 import，就把这一行删掉
    
    country = Counter()
    province_cn = Counter()
    
    for info in users.values():
        loc = (info.get('location') or '').strip()
        if not loc:
            continue
        parts = [p.strip() for p in loc.split('-') if p is not None]
        c = parts[0] if parts else ''
        if not c:
            c = '未知'
        country[c] += 1
    
        # 仅当国家为中国时，按省份统计（location 形如：中国-浙江省-杭州市）
        if c == '中国' and len(parts) >= 2:
            p = parts[1] or '未知省份'
            # 省份名称规范化：去掉常见后缀（省/市/自治区/特别行政区）及少数民族长后缀
            norm_map = {
                '内蒙古自治区': '内蒙古',
                '广西壮族自治区': '广西',
                '宁夏回族自治区': '宁夏',
                '新疆维吾尔自治区': '新疆',
                '香港特别行政区': '香港',
                '澳门特别行政区': '澳门',
            }
            p = norm_map.get(p, re.sub(r'(省|市|自治区|特别行政区)$', '', p))
            if not p:
                p = '未知省份'
            province_cn[p] += 1
    
    # 国家饼图：Top10 + 其它
    top_c = country.most_common(10)
    top_c_sum = sum(v for _, v in top_c)
    geo_pie = [{"name": n, "value": v} for n, v in top_c]
    other_c_sum = sum(country.values()) - top_c_sum
    if other_c_sum > 0:
        geo_pie.append({"name": "其它", "value": other_c_sum})
    
    # 中国省份饼图：Top10 + 其它（仅在有中国数据时返回）
    province_pie = []
    if province_cn:
        top_p = province_cn.most_common(10)
        top_p_sum = sum(v for _, v in top_p)
        province_pie = [{"name": n, "value": v} for n, v in top_p]
        other_p_sum = sum(province_cn.values()) - top_p_sum
        if other_p_sum > 0:
            province_pie.append({"name": "其它", "value": other_p_sum})
    
    return {
        "day_labels": day_labels, "day_sold": day_sold, "day_rev": day_rev, "day_new": day_new,
        "month_labels": mon_labels, "mon_sold": mon_sold, "mon_rev": mon_rev, "mon_new": mon_new,
        "kpis": kpis,
        "geo_pie": geo_pie,
        "province_pie": province_pie,  # ✅ 新增：给前端画“中国省份分布”
    }

# —— 管理员统计页 —— 
@app.route('/stats')
@admin_required
def stats():
    users = load_users()

    # 只保留“当前空间”里出现过的代理（owner）
    owners = {v.get('owner') for v in users.values() if v.get('owner')}

    # 只用属于这些代理的台账（role='agent'）
    ledger = [
        r for r in load_ledger()
        if r.get('role') == 'agent' and (not owners or r.get('admin') in owners)
    ]

    data = _aggregate_stats(users, ledger, agent_view=False)
    return render_template('stats.html', **data, agent_view=False)
#代理/分销商统计
@app.route('/sales/stats')
def sales_stats():
    """
    代理/分销商数据统计页面。
    用途：显示当前代理或分销商的数据统计。
    """
    # 检查权限：必须是代理或分销商
    if not session.get('agent') and not session.get('distributor'):
        return redirect(url_for('login'))
    
    users = load_users()
    # 获取当前用户名（代理或分销商）
    current_user = session.get('agent') or session.get('distributor')
    my_users = {k: v for k, v in users.items() if v.get('owner') == current_user}
    my_ledger = [r for r in load_ledger() if r.get('role')=='agent' and r.get('admin')==current_user]
    data = _aggregate_stats(my_users, my_ledger, agent_view=True)
    return render_template('stats.html', **data, agent_view=True)

@app.route('/sales/distributors')
@agent_required
def get_distributors():
    """
    获取当前代理名下的分销商列表（API接口）
    用于账号分配功能
    """
    current_user = session['agent']
    users = load_users()
    
    distributors = []
    for username, user_info in users.items():
        if user_info.get('is_distributor') and user_info.get('distributor_owner') == current_user:
            distributors.append({
                'username': username,
                'nickname': user_info.get('nickname', username),
                'enabled': user_info.get('enabled', True)
            })
    
    return jsonify({'code': 0, 'data': distributors})

@app.route('/sales/assign_accounts', methods=['POST'])
@agent_required
def assign_accounts():
    """
    账号分配处理接口
    将选中的账号分配给指定的分销商
    """
    current_user = session['agent']
    data = request.get_json()
    
    if not data or 'accounts' not in data or 'distributor' not in data:
        return jsonify({'success': False, 'message': '参数不完整'}), 400
    
    account_usernames = data['accounts']
    distributor_username = data['distributor']
    
    if not account_usernames or not distributor_username:
        return jsonify({'success': False, 'message': '请选择账号和分销商'}), 400
    
    users = load_users()
    
    # 验证分销商是否属于当前代理
    distributor_info = users.get(distributor_username)
    if not distributor_info or not distributor_info.get('is_distributor') or distributor_info.get('distributor_owner') != current_user:
        return jsonify({'success': False, 'message': '无效的分销商'}), 400
    
    # 验证并处理账号分配
    assigned_count = 0
    invalid_accounts = []
    
    for username in account_usernames:
        user_info = users.get(username)
        if not user_info:
            invalid_accounts.append(f'{username}: 用户不存在')
            continue

        state = user_info.get('accounting', {})
        if state.get('owner') != current_user:
            invalid_accounts.append(f'{username}: 不属于当前代理')
            continue
        if state.get('status') != ACCOUNT_STATUS_SOLD:
            invalid_accounts.append(f'{username}: 账号未售出')
            continue
        sale_type = state.get('sale_type')
        if sale_type not in (SALE_TYPE_DIRECT, None):
            invalid_accounts.append(f'{username}: 仅支持总部直销账户分配')
            continue
        current_manager = state.get('manager')
        if current_manager and current_manager not in {current_user, distributor_username}:
            invalid_accounts.append(f'{username}: 已由其他分销管理')
            continue

        update_account_state(
            user_info,
            manager=distributor_username,
            sale_type=sale_type or SALE_TYPE_DIRECT,
        )
        user_info['assigned_at'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        assigned_count += 1

    if assigned_count:
        save_users(users)

    if invalid_accounts:
        message = f'成功分配 {assigned_count} 个账号。以下账号无法分配：' + '; '.join(invalid_accounts)
    else:
        message = f'成功分配 {assigned_count} 个账号'
    
    return jsonify({
        'code': 0,
        'success': True, 
        'message': message,
        'assigned_count': assigned_count,
        'invalid_accounts': invalid_accounts
    })


@app.route('/distributor-management')
@agent_required
def distributor_management():
    """
    分销人员管理页面
    """
    current_user = session['agent']
    
    # 获取分页参数
    page = request.args.get('page', 1, type=int)
    per_page = 10  # 每页显示10个分销代理
    
    # 获取当前代理的所有分销代理
    all_distributors = []
    total_distributors = 0
    active_distributors = 0
    inactive_distributors = 0
    managed_users = 0
    
    users = load_users()
    
    for username, user_info in users.items():
        if user_info.get('is_distributor') and user_info.get('distributor_owner') == current_user:
            # 计算该分销代理管理的用户数量（通过assigned_distributor字段）
            managed_count = sum(
                1
                for u in users.values()
                if u.get('accounting', {}).get('manager') == username
                and u.get('accounting', {}).get('status') == ACCOUNT_STATUS_SOLD
            )
            
            all_distributors.append({
                'username': username,
                'nickname': user_info.get('nickname'),
                'enabled': user_info.get('enabled', True),
                'managed_count': managed_count,
                'created_at': user_info.get('created_at', '')
            })
            
            total_distributors += 1
            if user_info.get('enabled', True):
                active_distributors += 1
            else:
                inactive_distributors += 1
            managed_users += managed_count
    
    # 分页处理
    total_pages = (total_distributors + per_page - 1) // per_page if total_distributors > 0 else 1
    start_index = (page - 1) * per_page
    end_index = start_index + per_page
    distributors = all_distributors[start_index:end_index]
    
    # 分页信息
    pagination = {
        'page': page,
        'per_page': per_page,
        'total': total_distributors,
        'total_pages': total_pages,
        'has_prev': page > 1,
        'has_next': page < total_pages,
        'prev_page': page - 1 if page > 1 else None,
        'next_page': page + 1 if page < total_pages else None,
        'start_record': start_index + 1 if total_distributors > 0 else 0,
        'end_record': min(end_index, total_distributors)
    }
    
    # 获取可分配权限的用户（当前代理的普通用户）
    available_users = []
    for username, user_info in users.items():
        if (user_info.get('owner') == current_user and 
            not user_info.get('is_admin') and 
            not user_info.get('is_agent') and 
            not user_info.get('is_distributor')):
            available_users.append({
                'username': username,
                'nickname': user_info.get('nickname')
            })
    
    return render_template('distributor_management.html',
                         distributors=distributors,
                         total_distributors=total_distributors,
                         active_distributors=active_distributors,
                         inactive_distributors=inactive_distributors,
                         managed_users=managed_users,
                         available_users=available_users,
                         pagination=pagination)

@app.route('/assign-distributor', methods=['POST'])
@agent_required
def assign_distributor():
    """分配分销权限"""
    username = request.form.get('username')
    current_user = session['agent']
    
    users = load_users()
    
    if not username or username not in users:
        return jsonify({'success': False, 'message': '用户不存在'})
    
    user_info = users[username]
    
    # 检查是否是当前代理的用户
    if user_info.get('owner') != current_user:
        return jsonify({'success': False, 'message': '只能给自己的用户分配权限'})
    
    # 检查用户是否已经有其他权限
    if user_info.get('is_admin') or user_info.get('is_agent') or user_info.get('is_distributor'):
        return jsonify({'success': False, 'message': '该用户已有其他权限，无法分配分销权限'})
    
    # 分配分销权限
    users[username]['is_distributor'] = True
    users[username]['distributor_owner'] = current_user
    users[username]['created_at'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    save_users(users)
    return jsonify({'success': True, 'message': '分销权限分配成功'})

@app.route('/revoke-distributor', methods=['POST'])
@agent_required
def revoke_distributor():
    """撤销分销权限"""
    username = request.form.get('username')
    current_user = session['agent']
    
    users = load_users()
    
    if not username or username not in users:
        return jsonify({'success': False, 'message': '用户不存在'})
    
    user_info = users[username]
    
    # 检查是否是当前代理分配的分销权限
    if user_info.get('distributor_owner') != current_user:
        return jsonify({'success': False, 'message': '只能撤销自己分配的分销权限'})
    
    # 撤销分销权限
    users[username]['is_distributor'] = False
    users[username].pop('distributor_owner', None)
    
    save_users(users)
    return jsonify({'success': True, 'message': '分销权限撤销成功'})

@app.route('/api/distributor/<username>/managed-users')
@agent_required
def get_distributor_managed_users(username):
    """获取分销代理管理的用户列表"""
    current_user = session['agent']
    
    users = load_users()
    
    if username not in users:
        return jsonify({'success': False, 'message': '分销代理不存在'})
    
    distributor_info = users[username]
    
    # 检查是否是当前代理的分销代理
    if distributor_info.get('distributor_owner') != current_user:
        return jsonify({'success': False, 'message': '无权查看该分销代理的用户'})
    
    # 获取该分销代理管理的用户
    managed_users = []
    for user_username, user_info in users.items():
        state = user_info.get('accounting', {})
        if not state:
            continue
        if state.get('owner') == username or state.get('manager') == username:
            managed_users.append({
                'username': user_username,
                'nickname': user_info.get('nickname'),
                'enabled': user_info.get('enabled', True),
                'created_at': user_info.get('created_at', '')
            })
    
    return jsonify({'success': True, 'users': managed_users})


@app.route('/api/available-accounts-count')
@agent_required
def get_available_accounts_count():
    """
    获取当前代理的可用账户数量（真正的库存账户：forsale=True且未分配给分销商）
    """
    current_agent = session['agent']
    users = load_users()
    
    # 统计当前代理名下的真正库存账户数量
    available_count = sum(
        1
        for user_info in users.values()
        if user_info.get('accounting', {}).get('owner') == current_agent
        and user_info.get('accounting', {}).get('status') == ACCOUNT_STATUS_AGENT_STOCK
    )
    
    return jsonify({'success': True, 'count': available_count})


@app.route('/distributor/stock-apply', methods=['GET', 'POST'])
@distributor_required
def distributor_stock_apply():
    """
    分销商账号进货申请页面。
    用途：分销商提交进货申请，显示申请历史。
    交互：写入distribution_requests.json，session标记提交成功。
    """
    # 获取当前分销商绑定的代理商信息
    users = load_users()
    current_distributor = session.get('distributor')
    distributor_info = users.get(current_distributor, {})
    bound_agent = distributor_info.get('distributor_owner')
    
    products = load_products()
    if request.method == 'POST':
        quantity_str = request.form.get('quantity', '').strip()
        product = request.form.get('product', '').strip()
        agent = bound_agent  # 使用绑定的代理商
        
        # 验证必填字段
        if not quantity_str:
            session['error_message'] = '进货数量不能为空'
            return redirect(url_for('distributor_stock_apply'))
        
        if not product:
            session['error_message'] = '请选择产品类型'
            return redirect(url_for('distributor_stock_apply'))
        
        try:
            quantity = int(quantity_str)
            if quantity <= 0:
                session['error_message'] = '进货数量必须大于0'
                return redirect(url_for('distributor_stock_apply'))
            if quantity > 100:
                session['error_message'] = '进货数量不能超过100'
                return redirect(url_for('distributor_stock_apply'))
        except ValueError:
            session['error_message'] = '进货数量必须是有效数字'
            return redirect(url_for('distributor_stock_apply'))
        
        requests = load_distribution_requests()
        requests.append({
            'id': os.urandom(6).hex(),
            'distributor': session.get('distributor'),
            'agent': agent,
            'product': product,
            'quantity': quantity,
            'status': 'pending',
            'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        })
        save_distribution_requests(requests)
        session['stock_apply_success'] = True
        return redirect(url_for('distributor_stock_apply'))
    
    success = session.pop('stock_apply_success', False)
    error_message = session.pop('error_message', None)
    my_requests = [r for r in load_distribution_requests() if r.get('distributor') == session.get('distributor')]
    
    # 分页参数
    page = int(request.args.get('page', 1))
    per_page = 10  # 每页显示10条记录
    
    # 按时间倒序排序（最新的在前面）
    my_requests.sort(key=lambda x: x.get('created_at', ''), reverse=True)
    
    # 分页计算
    total_requests = len(my_requests)
    start_record = (page - 1) * per_page + 1
    end_record = min(page * per_page, total_requests)
    
    # 获取当前页的记录
    page_requests = my_requests[(page - 1) * per_page:page * per_page]
    
    # 分页信息
    pagination = {
        'page': page,
        'per_page': per_page,
        'total': total_requests,
        'total_pages': (total_requests - 1) // per_page + 1 if total_requests > 0 else 1,
        'has_prev': page > 1,
        'has_next': page < ((total_requests - 1) // per_page + 1),
        'prev_num': page - 1 if page > 1 else None,
        'next_num': page + 1 if page < ((total_requests - 1) // per_page + 1) else None,
        'start_record': start_record if total_requests > 0 else 0,
        'end_record': end_record if total_requests > 0 else 0
    }
    
    # 获取当前分销商绑定的代理商信息
    users = load_users()
    current_distributor = session.get('distributor')
    distributor_info = users.get(current_distributor, {})
    bound_agent = distributor_info.get('distributor_owner')
    
    # 获取绑定代理商的详细信息
    bound_agent_info = None
    if bound_agent and bound_agent in users:
        bound_agent_info = {
            'username': bound_agent,
            'nickname': users[bound_agent].get('nickname', bound_agent)
        }
    
    return render_template('distributor_stock_apply.html', 
                         products=products,
                         success=success,
                         error_message=error_message,
                         requests=page_requests,
                         pagination=pagination,
                         bound_agent=bound_agent_info)


@app.route('/distributor/users')
@distributor_required
def distributor_users():
    """
    分销商用户管理页面，显示分销商的账户。
    用途：分销商查看和管理自己的账户。
    """
    current_distributor = session['distributor']
    users = load_users()
    
    # 获取搜索和筛选参数
    search_username = request.args.get('q', '').strip().lower()
    search_nickname = request.args.get('nick', '').strip().lower()
    sale_filter = request.args.get('sale', '')
    start_date = request.args.get('start', '')
    end_date = request.args.get('end', '')
    sort_order = request.args.get('sort', 'desc')
    
    # 获取分页参数
    page = int(request.args.get('page', 1))
    per_page = 10  # 每页10个用户
    
    # 过滤出当前分销商的用户（库存或已托管）
    my_users = []
    for username, user_info in users.items():
        state = user_info.get('accounting', {})
        if not state:
            continue

        owns_inventory = (
            state.get('owner') == current_distributor
            and state.get('status') in {ACCOUNT_STATUS_DISTRIBUTOR_STOCK, ACCOUNT_STATUS_SOLD}
        )
        manages_account = (
            state.get('manager') == current_distributor
            and state.get('status') == ACCOUNT_STATUS_SOLD
        )

        if not owns_inventory and not manages_account:
            continue

        user_type = get_distributor_user_type(user_info, current_distributor)
        user_info['user_type'] = user_type

        if sale_filter == 'sold' and user_type != 'sold':
            continue
        if sale_filter == 'unsold' and user_type != 'unsold':
            continue

        if search_username and search_username not in username.lower():
            continue
        if search_nickname and search_nickname not in user_info.get('nickname', '').lower():
            continue

        created_at = user_info.get('created_at', '')
        if start_date and created_at and created_at < start_date:
            continue
        if end_date and created_at and created_at > end_date:
            continue

        my_users.append({
            'username': username,
            **user_info
        })
    
    # 排序：优先显示未售账户，然后按创建时间排序
    def sort_key(user):
        is_unsold = user.get('user_type') == 'unsold'
        created_at = user.get('created_at', '')
        return (not is_unsold, created_at)
    
    if sort_order == 'asc':
        my_users.sort(key=sort_key)
    else:
        my_users.sort(key=sort_key, reverse=False)  # 保持未售优先，但时间可以倒序
    
    # 统计数据
    total_users = len(my_users)
    unsold_users = sum(1 for user in my_users if user.get('user_type') == 'unsold')
    sold_users = sum(1 for user in my_users if user.get('user_type') == 'sold')
    managed_users = sum(1 for user in my_users if user.get('user_type') == 'managed')
    
    # 分页计算
    total_pages = (total_users + per_page - 1) // per_page  # 向上取整
    start_index = (page - 1) * per_page
    end_index = start_index + per_page
    paginated_users = my_users[start_index:end_index]
    
    # 分页信息（使用与ledger_view相同的结构）
    pagination = {
        'page': page,
        'per_page': per_page,
        'total': total_users,
        'total_pages': total_pages,
        'has_prev': page > 1,
        'has_next': page < total_pages,
        'prev_page': page - 1 if page > 1 else None,
        'next_page': page + 1 if page < total_pages else None,
        'start_record': start_index + 1 if total_users > 0 else 0,
        'end_record': min(end_index, total_users)
    }
    
    return render_template('distributor_users.html',
                         users=paginated_users,
                         total_users=total_users,
                         unsold_users=unsold_users,
                         sold_users=sold_users,
                         managed_users=managed_users,
                         pagination=pagination)


def get_distributor_user_type(user_info, distributor):
    """获取分销商视角下的用户类型。"""

    state = user_info.get('accounting', {})
    if not state:
        return 'unknown'

    if state.get('owner') == distributor and state.get('status') == ACCOUNT_STATUS_DISTRIBUTOR_STOCK:
        return 'unsold'

    if state.get('status') == ACCOUNT_STATUS_SOLD and state.get('manager') == distributor:
        if state.get('sale_type') == SALE_TYPE_DISTRIBUTION and state.get('owner') == distributor:
            return 'sold'
        return 'managed'

    return 'unknown'


@app.route('/distributor/mark-sold', methods=['POST'])
@distributor_required
def distributor_mark_sold():
    """
    分销商标记单个用户为已售
    """
    data = request.get_json()
    username = data.get('username')
    
    users = load_users()
    current_distributor = session['distributor']
    
    if username in users:
        state = users[username].get('accounting', {})
        if state.get('owner') == current_distributor and state.get('status') == ACCOUNT_STATUS_DISTRIBUTOR_STOCK:
            sold_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            update_account_state(
                users[username],
                owner=current_distributor,
                manager=current_distributor,
                status=ACCOUNT_STATUS_SOLD,
                sale_type=SALE_TYPE_DISTRIBUTION,
                sold_at=sold_time,
            )
            save_users(users)

            records = load_ledger()
            record_transaction(
                records,
                transaction_type=TRANSACTION_DISTRIBUTOR_SALE,
                actor=current_distributor,
                actor_role=ROLE_DISTRIBUTOR,
                amount=float(users[username].get('price', 0) or 0),
                quantity=1,
                product=users[username].get('product', '未知产品'),
                account_username=username,
                sale_type=SALE_TYPE_DISTRIBUTION,
            )
            save_ledger(records)

        return jsonify({'success': True})
    
    return jsonify({'success': False}), 400


@app.route('/distributor/mark-unsold', methods=['POST'])
@distributor_required
def distributor_mark_unsold():
    """
    分销商标记单个用户为未售
    注意：只有未记录到台账的用户才能标记为未售
    """
    data = request.get_json()
    username = data.get('username')
    
    users = load_users()
    current_distributor = session['distributor']
    
    if username in users:
        state = users[username].get('accounting', {})
        if (
            state.get('owner') == current_distributor
            and state.get('status') == ACCOUNT_STATUS_SOLD
            and state.get('sale_type') == SALE_TYPE_DISTRIBUTION
        ):
            update_account_state(
                users[username],
                owner=current_distributor,
                manager=current_distributor,
                status=ACCOUNT_STATUS_DISTRIBUTOR_STOCK,
                sale_type=None,
                sold_at=None,
            )
            save_users(users)
            return jsonify({'success': True})

    return jsonify({'success': False}), 400


@app.route('/distributor/export-users', methods=['POST'])
@distributor_required
def distributor_export_users():
    """
    分销商导出选中用户
    """
    names = request.form.getlist('names')
    users = load_users()
    current_distributor = session['distributor']
    
    # 过滤出当前分销商的用户
    selected_users = {}
    for name in names:
        state = users.get(name, {}).get('accounting', {})
        if not state:
            continue
        owns_inventory = (
            state.get('owner') == current_distributor
            and state.get('status') in {ACCOUNT_STATUS_DISTRIBUTOR_STOCK, ACCOUNT_STATUS_SOLD}
        )
        manages_account = (
            state.get('manager') == current_distributor
            and state.get('status') == ACCOUNT_STATUS_SOLD
        )
        if owns_inventory or manages_account:
            selected_users[name] = users[name]
    
    if not selected_users:
        return redirect(url_for('distributor_users'))
    
    # 创建Excel文件
    wb = Workbook()
    ws = wb.active
    ws.title = "分销用户信息"
    
    # 设置表头
    ws.append(['用户名', '密码', '昵称', '产品', '分销状态', '创建时间', '备注'])
    
    # 添加用户数据
    for username, info in selected_users.items():
        user_type = get_distributor_user_type(info, current_distributor)
        ws.append([
            username,
            info.get('password', ''),
            info.get('nickname', ''),
            info.get('product', ''),
            {
                'sold': '已售出',
                'unsold': '库存',
                'managed': '托管客户',
            }.get(user_type, '未知'),
            info.get('created_at', ''),
            info.get('remark', '')
        ])
    
    # 生成文件
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    
    # 生成文件名（包含分销商名和时间戳）
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'{current_distributor}_分销用户_{timestamp}.xlsx'
    
    return send_file(bio, download_name=filename, as_attachment=True)


@app.route('/distributor/ledger')
@distributor_required
def distributor_ledger():
    """
    分销商销售台账页面。
    用途：显示当前分销商的销售记录与统计。
    """
    # 获取当前分销商用户名
    current_distributor = session.get('distributor')
    
    product_filter = request.args.get('product', '')
    start = request.args.get('start', '')
    end = request.args.get('end', '')
    # 获取分页参数
    page = int(request.args.get('page', 1))
    per_page = 10  # 每页显示10条记录

    # 筛选出当前分销商的销售记录
    records = [
        r for r in load_ledger()
        if (r.get('actor') or r.get('admin')) == current_distributor
        and (r.get('actor_role') or r.get('role')) == 'distributor'
    ]
    if product_filter:
        records = [r for r in records if r.get('product') == product_filter]
    
    from datetime import datetime
    today = datetime.now().strftime('%Y-%m-%d')
    this_month = datetime.now().strftime('%Y-%m')
    this_year = datetime.now().strftime('%Y')

    # 应用日期筛选
    filtered_records = []
    for r in records:
        record_time = r.get('time', '')
        
        if start and record_time < start:
            continue
        if end and record_time > end:
            continue
        
        filtered_records.append(r)

    # 按时间倒序排序
    filtered_records.sort(key=lambda x: x.get('time', ''), reverse=True)

    # 统计计算
    daily = monthly = yearly = total = 0.0

    for r in filtered_records:
        if r.get('direction', 'in') != 'in':
            continue
        amount = float(r.get('amount', r.get('revenue', 0)) or 0)
        total += amount
        record_time = r.get('time', '')

        if record_time.startswith(today):
            daily += amount
        if record_time.startswith(this_month):
            monthly += amount
        if record_time.startswith(this_year):
            yearly += amount

    # 实现分页
    total_records = len(filtered_records)
    start_index = (page - 1) * per_page
    end_index = start_index + per_page
    paginated_records = filtered_records[start_index:end_index]
    
    # 计算总页数
    total_pages = max(1, (total_records + per_page - 1) // per_page)

    products = load_products()
    return render_template(
        'agent_ledger.html', 
        records=paginated_records,
        product_filter=product_filter,
        start=start, end=end, products=products,
        daily=daily, monthly=monthly, yearly=yearly, total=total,
        page=page, per_page=per_page, total_pages=total_pages, total_records=total_records,
        start_index=start_index, end_index=end_index
    )


@app.route('/update-user-field', methods=['POST'])
@distributor_required
def update_user_field():
    """
    更新用户字段（昵称、备注等）
    """
    data = request.get_json()
    username = data.get('username')
    field = data.get('field')
    value = data.get('value')
    
    if not username or not field:
        return jsonify({'success': False, 'message': '参数不完整'}), 400
    
    users = load_users()
    current_distributor = session['distributor']
    
    # 检查用户是否存在且属于当前分销商
    if username not in users:
        return jsonify({'success': False, 'message': '用户不存在'}), 404
    
    state = users[username].get('accounting', {})
    if state.get('owner') != current_distributor and state.get('manager') != current_distributor:
        return jsonify({'success': False, 'message': '无权限修改此用户'}), 403
    
    # 更新字段
    if field in ['nickname', 'remark']:
        users[username][field] = value
        save_users(users)
        return jsonify({'success': True})
    else:
        return jsonify({'success': False, 'message': '不支持的字段'}), 400


@app.route('/distribution-approval')
@agent_required
def distribution_approval():
    """
    分销进货审批页面，显示所有分销进货申请。
    用途：代理角色审批分销进货申请。
    """
    current_agent = session['agent']
    requests = load_distribution_requests()
    
    # 过滤出当前代理的申请
    agent_requests = [req for req in requests if req.get('agent') == current_agent]
    
    # 统计数据
    pending_count = sum(1 for req in agent_requests if req.get('status') == 'pending')
    approved_count = sum(1 for req in agent_requests if req.get('status') == 'approved')
    rejected_count = sum(1 for req in agent_requests if req.get('status') == 'rejected')
    
    # 计算可用账户数量（真正的库存账户：forsale=True且未分配给分销商）
    users = load_users()
    available_accounts = sum(
        1
        for user in users.values()
        if user.get('accounting', {}).get('owner') == current_agent
        and user.get('accounting', {}).get('status') == ACCOUNT_STATUS_AGENT_STOCK
    )
    
    return render_template('distribution_approval.html', 
                         requests=agent_requests,
                         pending_count=pending_count,
                         approved_count=approved_count,
                         rejected_count=rejected_count,
                         available_accounts=available_accounts)


@app.route('/distribution-approval/<request_id>/approve', methods=['POST'])
@agent_required
def approve_distribution_request(request_id):
    """
    审批通过分销进货申请。
    用途：代理角色操作。
    """
    current_agent = session['agent']
    requests = load_distribution_requests()
    
    # 查找申请记录
    request_record = next((req for req in requests if req['id'] == request_id), None)
    if not request_record or request_record.get('agent') != current_agent:
        return jsonify({'success': False, 'message': '申请不存在或无权操作'})
    
    if request_record.get('status') != 'pending':
        return jsonify({'success': False, 'message': '申请状态不允许操作'})
    
    # 检查库存
    users = load_users()
    available_accounts = [
        username
        for username, user in users.items()
        if user.get('accounting', {}).get('owner') == current_agent
        and user.get('accounting', {}).get('status') == ACCOUNT_STATUS_AGENT_STOCK
    ]
    
    requested_quantity = request_record.get('quantity', 0)
    
    if len(available_accounts) < requested_quantity:
        shortage = requested_quantity - len(available_accounts)
        return jsonify({
            'success': False, 
            'message': f'存货不足，请先进行采购，差额：{shortage}个'
        })
    
    # 随机选择账户分配给分销角色
    selected_accounts = random.sample(available_accounts, requested_quantity)
    distributor = request_record.get('distributor')
    
    sale_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    for username in selected_accounts:
        update_account_state(
            users[username],
            owner=distributor,
            manager=distributor,
            status=ACCOUNT_STATUS_DISTRIBUTOR_STOCK,
            sale_type=None,
            sold_at=None,
        )
        users[username]['transferred_at'] = sale_time

    save_users(users)
    
    # 更新申请状态
    request_record['status'] = 'approved'
    request_record['approved_at'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    request_record['assigned_accounts'] = selected_accounts
    
    save_distribution_requests(requests)
    
    # 添加台账记录 - 记录分销售出
    records = load_ledger()
    product_name = request_record.get('product', '未知产品')

    products = load_products()
    product_price = float(products.get(product_name, {}).get('price', 0) or 0)
    total_amount = product_price * requested_quantity

    record_transaction(
        records,
        transaction_type=TRANSACTION_AGENT_TO_DISTRIBUTOR,
        actor=current_agent,
        actor_role=ROLE_AGENT,
        amount=total_amount,
        quantity=requested_quantity,
        product=product_name,
        counterparty=distributor,
    )
    record_transaction(
        records,
        transaction_type=TRANSACTION_AGENT_TO_DISTRIBUTOR,
        actor=distributor,
        actor_role=ROLE_DISTRIBUTOR,
        amount=total_amount,
        quantity=requested_quantity,
        product=product_name,
        counterparty=current_agent,
        direction='out',
    )

    save_ledger(records)
    
    return jsonify({'success': True, 'message': '审批通过，账户已分配'})


@app.route('/distribution-approval/<request_id>/reject', methods=['POST'])
@agent_required
def reject_distribution_request(request_id):
    """
    拒绝分销进货申请。
    用途：代理角色操作。
    """
    current_agent = session['agent']
    requests = load_distribution_requests()
    
    # 查找申请记录
    request_record = next((req for req in requests if req['id'] == request_id), None)
    if not request_record or request_record.get('agent') != current_agent:
        return jsonify({'success': False, 'message': '申请不存在或无权操作'})
    
    if request_record.get('status') != 'pending':
        return jsonify({'success': False, 'message': '申请状态不允许操作'})
    
    # 更新申请状态
    request_record['status'] = 'rejected'
    request_record['rejected_at'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    save_distribution_requests(requests)
    
    return jsonify({'success': True, 'message': '申请已拒绝'})


if __name__ == '__main__':
    # 主程序入口，支持命令行指定端口
    parser = argparse.ArgumentParser()
    parser.add_argument('--port', type=int, default=5001, help='Port to run the server on')
    args = parser.parse_args()
    app.run(host='0.0.0.0', port=args.port, debug=True)

